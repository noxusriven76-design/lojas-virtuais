from __future__ import annotations

import re
import csv
import io
from datetime import datetime
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.core.audit import write_audit_log
from app.core.deps import get_current_user, get_db
from app.core.permissions import require_catalog_write
from app.core.uploads import UploadValidationError, delete_upload_by_public_url, get_product_upload_dir, save_upload_file
from app.models.catalog_job import CatalogJob
from app.models.catalog import Category, Product, ProductImage, ProductVariant
from app.models.order import OrderItem
from app.schemas.catalog import (
    CatalogImportOut,
    CatalogJobOut,
    CategoryCreateIn,
    CategoryOut,
    CategoryTreeOut,
    CategoryUpdateIn,
    ProductBulkUpdateIn,
    ProductBulkUpdateOut,
    ProductAdminOut,
    ProductCreateIn,
    ProductDeleteOut,
    ProductImageOut,
    ProductImageAdminOut,
    ProductImageUpdateIn,
    ProductVariantCreateIn,
    ProductUpdateIn,
)
from app.schemas.user import UserOut

router = APIRouter(prefix="/admin/stores/{store_id}")


def _ensure_store_category(db: Session, store_id: int, category_id: int) -> Category:
    category = db.query(Category).filter(Category.store_id == store_id, Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="category not found")
    return category


def _check_parent_cycle(db: Session, store_id: int, category_id: int, parent_id: int) -> None:
    if parent_id == category_id:
        raise HTTPException(status_code=400, detail="parent_id cannot reference the category itself")

    if not hasattr(Category, "parent_id"):
        raise HTTPException(status_code=400, detail="parent_id is not supported for categories")

    visited: set[int] = set()
    current_parent_id = parent_id
    while current_parent_id is not None:
        if current_parent_id == category_id:
            raise HTTPException(status_code=400, detail="parent_id cannot reference a descendant category")
        if current_parent_id in visited:
            # Existing corrupted graph: stop recursion to avoid infinite loop.
            break
        visited.add(current_parent_id)

        row = (
            db.query(Category.parent_id)
            .filter(Category.store_id == store_id, Category.id == current_parent_id)
            .first()
        )
        current_parent_id = row[0] if row else None


def _ensure_store_product(db: Session, store_id: int, product_id: int) -> Product:
    product = db.query(Product).filter(Product.store_id == store_id, Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="product not found")
    return product


def _ensure_store_product_image(db: Session, store_id: int, product_id: int, image_id: int) -> ProductImage:
    image = (
        db.query(ProductImage)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id, ProductImage.id == image_id)
        .first()
    )
    if not image:
        raise HTTPException(status_code=404, detail="product image not found")
    return image


def _sync_product_cover_from_gallery(db: Session, store_id: int, product: Product) -> None:
    cover = (
        db.query(ProductImage)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == product.id, ProductImage.is_cover == True)  # noqa: E712
        .order_by(ProductImage.sort_order.asc(), ProductImage.id.asc())
        .first()
    )
    product.image_url = cover.image_url if cover else None
    db.add(product)


def _set_cover_image(db: Session, store_id: int, product_id: int, image_id: int) -> None:
    images = db.query(ProductImage).filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id).all()
    for image in images:
        image.is_cover = image.id == image_id
        db.add(image)


def _split_unique_csv(value: str | None, fallback: str) -> list[str]:
    raw = str(value or "").strip()
    parts = [item.strip() for item in raw.split(",")] if raw else []
    cleaned = [item for item in parts if item]
    if not cleaned:
        cleaned = [fallback]
    out: list[str] = []
    seen: set[str] = set()
    for item in cleaned:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def _slug_token(value: str) -> str:
    token = re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")
    return token[:20] or "var"


def _join_unique(values: list[str], fallback: str) -> str:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        item = str(value or "").strip()
        if not item:
            continue
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    if not out:
        return fallback
    return ", ".join(out)


def _build_unique_sku(
    db: Session,
    *,
    store_id: int,
    candidate: str,
    current_variant_id: int | None = None,
) -> str:
    base = (candidate or "").strip() or "SKU"
    sku = base
    idx = 2
    while True:
        exists = (
            db.query(ProductVariant.id)
            .filter(ProductVariant.store_id == store_id, ProductVariant.sku == sku)
            .first()
        )
        if not exists or (current_variant_id is not None and exists[0] == current_variant_id):
            return sku
        sku = f"{base}-{idx}"
        idx += 1


def _snapshot_product(product: Product) -> dict:
    return {
        "id": product.id,
        "store_id": product.store_id,
        "category_id": product.category_id,
        "name": product.name,
        "description": product.description or "",
        "image_url": product.image_url,
        "base_price": float(product.base_price),
        "is_active": bool(product.is_active),
    }


def _snapshot_primary_variant(db: Session, *, store_id: int, product_id: int) -> dict | None:
    variant = (
        db.query(ProductVariant)
        .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product_id)
        .order_by(ProductVariant.id.asc())
        .first()
    )
    if not variant:
        return None
    return {
        "id": variant.id,
        "sku": variant.sku,
        "color": variant.color,
        "size": variant.size,
        "price": float(variant.price),
        "stock": int(variant.stock),
        "active": bool(variant.active),
    }


def _validate_product_values(
    *,
    name: str,
    description: str,
    base_price: Decimal,
    sku: str | None,
    color: str | None,
    size: str | None,
    stock: int | None,
) -> None:
    clean_name = (name or "").strip()
    if len(clean_name) < 2:
        raise HTTPException(status_code=400, detail="name must have at least 2 characters")
    if len(clean_name) > 180:
        raise HTTPException(status_code=400, detail="name too long (max 180)")
    if len(description or "") > 2000:
        raise HTTPException(status_code=400, detail="description too long (max 2000)")
    if base_price <= 0:
        raise HTTPException(status_code=400, detail="base_price must be greater than 0")
    if sku and len(sku.strip()) > 80:
        raise HTTPException(status_code=400, detail="sku too long (max 80)")
    if color and len(color) > 200:
        raise HTTPException(status_code=400, detail="color list too long (max 200)")
    if size and len(size) > 200:
        raise HTTPException(status_code=400, detail="size list too long (max 200)")
    if stock is not None and int(stock) < 0:
        raise HTTPException(status_code=400, detail="stock must be greater than or equal to 0")


def _upsert_primary_variant(
    db: Session,
    *,
    store_id: int,
    product: Product,
    color: str | None,
    size: str | None,
    stock: int | None,
    sku: str | None,
    force_price_sync: bool = True,
) -> ProductVariant:
    variant = (
        db.query(ProductVariant)
        .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product.id)
        .order_by(ProductVariant.id.asc())
        .first()
    )
    color_input = None if color is None else (color or "").strip()
    size_input = None if size is None else (size or "").strip()
    stock_value = int(stock) if stock is not None else (variant.stock if variant else 0)
    if stock_value < 0:
        raise HTTPException(status_code=400, detail="stock must be greater than or equal to 0")
    sku_candidate = (sku or "").strip() or (variant.sku if variant else f"SKU-{product.id}")

    if variant is None:
        color_value = color_input or "Padrao"
        size_value = size_input or "Unico"
        variant = ProductVariant(
            store_id=store_id,
            product_id=product.id,
            sku="",
            color=color_value,
            size=size_value,
            price=product.base_price,
            stock=stock_value,
            active=True,
        )
        db.add(variant)
        db.flush()
    else:
        variant.color = (
            (color_input or "Padrao")
            if color_input is not None
            else (variant.color or "Padrao")
        )
        variant.size = (
            (size_input or "Unico")
            if size_input is not None
            else (variant.size or "Unico")
        )
        variant.stock = stock_value
        variant.active = True
        if force_price_sync:
            variant.price = product.base_price

    variant.sku = _build_unique_sku(
        db,
        store_id=store_id,
        candidate=sku_candidate,
        current_variant_id=variant.id,
    )
    db.add(variant)
    return variant


def _sync_product_variants(
    db: Session,
    *,
    store_id: int,
    product: Product,
    colors_raw: str | None,
    sizes_raw: str | None,
    stock: int | None,
    sku: str | None,
    sync_all_prices: bool = False,
) -> None:
    current_variants = (
        db.query(ProductVariant)
        .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product.id)
        .order_by(ProductVariant.id.asc())
        .all()
    )
    primary_current = current_variants[0] if current_variants else None
    fallback_color = primary_current.color if primary_current and primary_current.color else "Padrao"
    fallback_size = primary_current.size if primary_current and primary_current.size else "Unico"
    colors = _split_unique_csv(colors_raw, fallback_color)
    sizes = _split_unique_csv(sizes_raw, fallback_size)

    stock_value = int(stock) if stock is not None else (primary_current.stock if primary_current else 0)
    if stock_value < 0:
        raise HTTPException(status_code=400, detail="stock must be greater than or equal to 0")

    sku_base = (sku or "").strip() or (primary_current.sku if primary_current else f"SKU-{product.id}")

    primary_variant = _upsert_primary_variant(
        db,
        store_id=store_id,
        product=product,
        color=colors[0],
        size=sizes[0],
        stock=stock_value,
        sku=sku_base,
        force_price_sync=True,
    )

    existing_by_combo: dict[tuple[str, str], ProductVariant] = {}
    for variant in (
        db.query(ProductVariant)
        .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product.id)
        .all()
    ):
        combo = ((variant.color or "").strip().lower(), (variant.size or "").strip().lower())
        existing_by_combo[combo] = variant

    first_combo = (colors[0].strip().lower(), sizes[0].strip().lower())
    for color in colors:
        for size in sizes:
            combo = (color.strip().lower(), size.strip().lower())
            if combo == first_combo:
                continue
            variant = existing_by_combo.get(combo)
            if variant is None:
                sku_candidate = f"{sku_base}-{_slug_token(color)}-{_slug_token(size)}"
                variant = ProductVariant(
                    store_id=store_id,
                    product_id=product.id,
                    sku="",
                    color=color,
                    size=size,
                    price=product.base_price,
                    stock=stock_value,
                    active=True,
                )
                db.add(variant)
                db.flush()
                variant.sku = _build_unique_sku(
                    db,
                    store_id=store_id,
                    candidate=sku_candidate,
                    current_variant_id=variant.id,
                )
            else:
                variant.color = color
                variant.size = size
                if stock is not None:
                    variant.stock = stock_value
                if sync_all_prices:
                    variant.price = product.base_price
                variant.active = True
            db.add(variant)

    if sync_all_prices:
        all_variants = (
            db.query(ProductVariant)
            .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product.id)
            .all()
        )
        for variant in all_variants:
            variant.price = product.base_price
            variant.active = True
            db.add(variant)

    db.add(primary_variant)


@router.get("/categories", response_model=list[CategoryTreeOut], response_model_exclude_none=True)
def list_categories(
    store_id: int,
    tree: bool = Query(default=False),
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    query = db.query(Category).filter(Category.store_id == store_id)
    if hasattr(Category, "sort_order"):
        rows = query.order_by(Category.sort_order.asc(), Category.name.asc()).all()
    else:
        rows = query.order_by(Category.name.asc()).all()

    if not tree:
        return [CategoryTreeOut.model_validate(r) for r in rows]

    nodes: dict[int, CategoryTreeOut] = {}
    for row in rows:
        node = CategoryTreeOut.model_validate(row)
        node.children = []
        nodes[row.id] = node

    roots: list[CategoryTreeOut] = []
    for row in rows:
        node = nodes[row.id]
        parent_id = getattr(row, "parent_id", None)
        if parent_id is None:
            roots.append(node)
            continue
        parent = nodes.get(parent_id)
        if parent is None:
            roots.append(node)
            continue
        parent.children.append(node)

    return roots


@router.post("/categories", response_model=CategoryOut)
def create_category(
    store_id: int,
    payload: CategoryCreateIn,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    c = Category(store_id=store_id, name=name)
    db.add(c)
    db.commit()
    db.refresh(c)
    return CategoryOut.model_validate(c)


@router.patch("/categories/{category_id}", response_model=CategoryOut)
def update_category(
    store_id: int,
    category_id: int,
    payload: CategoryUpdateIn,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    category = _ensure_store_category(db, store_id=store_id, category_id=category_id)
    fields_set = payload.model_fields_set

    if "name" in fields_set:
        if payload.name is None:
            raise HTTPException(status_code=400, detail="name cannot be null")
        category.name = payload.name

    if "parent_id" in fields_set:
        if payload.parent_id is not None:
            parent = _ensure_store_category(db, store_id=store_id, category_id=payload.parent_id)
            _check_parent_cycle(db, store_id=store_id, category_id=category_id, parent_id=parent.id)
        if hasattr(category, "parent_id"):
            category.parent_id = payload.parent_id
        elif payload.parent_id is not None:
            raise HTTPException(status_code=400, detail="parent_id is not supported for categories")

    if "is_active" in fields_set:
        if hasattr(category, "is_active"):
            if payload.is_active is None:
                raise HTTPException(status_code=400, detail="is_active cannot be null")
            category.is_active = payload.is_active
        elif payload.is_active is not None:
            raise HTTPException(status_code=400, detail="is_active is not supported for categories")

    if "sort_order" in fields_set:
        if hasattr(category, "sort_order"):
            if payload.sort_order is None:
                raise HTTPException(status_code=400, detail="sort_order cannot be null")
            category.sort_order = payload.sort_order
        elif payload.sort_order is not None:
            raise HTTPException(status_code=400, detail="sort_order is not supported for categories")

    db.add(category)
    db.commit()
    db.refresh(category)
    return CategoryOut.model_validate(category)


@router.delete("/categories/{category_id}")
def delete_category(
    store_id: int,
    category_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    category = _ensure_store_category(db, store_id=store_id, category_id=category_id)

    if hasattr(Category, "parent_id"):
        has_children = (
            db.query(Category.id)
            .filter(Category.store_id == store_id, Category.parent_id == category_id)
            .first()
            is not None
        )
        if has_children:
            raise HTTPException(status_code=409, detail="Cannot delete category with subcategories")

    has_products = (
        db.query(Product.id).filter(Product.store_id == store_id, Product.category_id == category_id).first() is not None
    )
    if has_products:
        raise HTTPException(status_code=409, detail="Cannot delete category with linked products")

    db.delete(category)
    db.commit()
    return {"ok": True}


@router.get("/products")
def list_products(store_id: int, db: Session = Depends(get_db), _=Depends(require_catalog_write)):
    rows = db.query(Product).filter(Product.store_id == store_id).order_by(Product.id.desc()).all()
    output = []
    for p in rows:
        variants = (
            db.query(ProductVariant)
            .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == p.id)
            .order_by(ProductVariant.id.asc())
            .all()
        )
        primary_variant = variants[0] if variants else None
        colors_text = _join_unique([v.color for v in variants], "Padrao")
        sizes_text = _join_unique([v.size for v in variants], "Unico")
        output.append(
            {
                "id": p.id,
                "category_id": p.category_id,
                "name": p.name,
                "description": p.description or "",
                "image_url": p.image_url,
                "images": [
                    ProductImageAdminOut.model_validate(image).model_dump()
                    for image in (
                        db.query(ProductImage)
                        .filter(ProductImage.store_id == store_id, ProductImage.product_id == p.id)
                        .order_by(ProductImage.sort_order.asc(), ProductImage.id.asc())
                        .all()
                    )
                ],
                "base_price": p.base_price,
                "is_active": p.is_active,
                "sku": primary_variant.sku if primary_variant else "",
                "color": colors_text,
                "size": sizes_text,
                "stock": int(primary_variant.stock) if primary_variant else 0,
            }
        )
    return output


@router.get("/products/export.csv")
def export_products_csv(
    store_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    rows = db.query(Product).filter(Product.store_id == store_id).order_by(Product.id.asc()).all()
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["id", "category_id", "name", "description", "base_price", "is_active", "sku", "color", "size", "stock"])
    for product in rows:
        primary = (
            db.query(ProductVariant)
            .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product.id)
            .order_by(ProductVariant.id.asc())
            .first()
        )
        writer.writerow(
            [
                product.id,
                product.category_id,
                product.name,
                product.description or "",
                float(product.base_price),
                int(product.is_active),
                primary.sku if primary else "",
                primary.color if primary else "",
                primary.size if primary else "",
                int(primary.stock) if primary else 0,
            ]
        )
    filename = f"catalog_store_{store_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _process_import_rows(
    db: Session,
    *,
    store_id: int,
    rows: list[dict[str, str]],
) -> tuple[int, int, list[str]]:
    imported = 0
    updated = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):
        try:
            name = str(row.get("name") or "").strip()
            category_id = int(row.get("category_id") or 0)
            base_price = Decimal(str(row.get("base_price") or "0"))
            sku = str(row.get("sku") or "").strip()
            description = str(row.get("description") or "")
            color = str(row.get("color") or "Padrao")
            size = str(row.get("size") or "Unico")
            stock = int(row.get("stock") or 0)
            is_active = str(row.get("is_active") or "1").strip() in {"1", "true", "True", "sim", "yes"}

            _ensure_store_category(db, store_id=store_id, category_id=category_id)
            _validate_product_values(
                name=name,
                description=description,
                base_price=base_price,
                sku=sku or None,
                color=color,
                size=size,
                stock=stock,
            )

            existing_variant = None
            if sku:
                existing_variant = (
                    db.query(ProductVariant)
                    .filter(ProductVariant.store_id == store_id, ProductVariant.sku == sku)
                    .first()
                )

            if existing_variant:
                product = _ensure_store_product(db, store_id=store_id, product_id=existing_variant.product_id)
                product.category_id = category_id
                product.name = name
                product.description = description
                product.base_price = base_price
                product.is_active = is_active
                _sync_product_variants(
                    db,
                    store_id=store_id,
                    product=product,
                    colors_raw=color,
                    sizes_raw=size,
                    stock=stock,
                    sku=sku,
                    sync_all_prices=True,
                )
                db.add(product)
                updated += 1
            else:
                product = Product(
                    store_id=store_id,
                    category_id=category_id,
                    name=name,
                    description=description,
                    image_url="",
                    base_price=base_price,
                    is_active=is_active,
                )
                db.add(product)
                db.flush()
                _sync_product_variants(
                    db,
                    store_id=store_id,
                    product=product,
                    colors_raw=color,
                    sizes_raw=size,
                    stock=stock,
                    sku=sku or None,
                )
                imported += 1
        except Exception as exc:
            errors.append(f"linha {idx}: {exc}")

    return imported, updated, errors


@router.post("/products/import", response_model=CatalogImportOut)
async def import_products(
    store_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
    user: UserOut = Depends(get_current_user),
):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only .csv or .xlsx files are allowed")

    job = CatalogJob(
        store_id=store_id,
        user_id=user.id,
        job_type="product_import",
        status="pending",
        payload={"filename": file.filename},
    )
    db.add(job)
    db.flush()
    job.status = "running"
    job.started_at = datetime.utcnow()
    db.add(job)
    db.commit()

    rows: list[dict[str, str]] = []
    if filename.endswith(".csv"):
        raw = await file.read()
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(row) for row in reader]
    else:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(status_code=400, detail="xlsx import requires openpyxl")
        raw = await file.read()
        wb = load_workbook(filename=io.BytesIO(raw), read_only=True)
        ws = wb.active
        header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for excel_row in ws.iter_rows(min_row=2):
            record: dict[str, str] = {}
            for idx, cell in enumerate(excel_row):
                key = header[idx] if idx < len(header) else f"col_{idx}"
                record[key] = "" if cell.value is None else str(cell.value)
            rows.append(record)

    imported, updated, errors = _process_import_rows(db, store_id=store_id, rows=rows)
    job = db.query(CatalogJob).filter(CatalogJob.store_id == store_id, CatalogJob.id == job.id).first()
    job.status = "completed" if not errors else "completed_with_errors"
    job.finished_at = datetime.utcnow()
    job.result = {"imported": imported, "updated": updated, "errors_count": len(errors)}
    job.error_message = "\n".join(errors[:10]) if errors else None
    db.add(job)
    db.commit()

    write_audit_log(
        db,
        request=request,
        user=user,
        store_id=store_id,
        action="catalog.import",
        entity_type="catalog_job",
        entity_id=job.id,
        before_data=None,
        after_data={"imported": imported, "updated": updated, "errors_count": len(errors)},
    )
    db.commit()

    return CatalogImportOut(ok=True, job_id=job.id, imported=imported, updated=updated, errors=errors[:50])


@router.post("/products")
def create_product(
    store_id: int,
    payload: ProductCreateIn,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
    user: UserOut = Depends(get_current_user),
):
    required = ["category_id", "name", "base_price"]
    for k in required:
        if getattr(payload, k) in (None, ""):
            raise HTTPException(status_code=400, detail=f"{k} required")

    base_price_value = Decimal(str(payload.base_price))
    _ensure_store_category(db, store_id=store_id, category_id=int(payload.category_id))
    _validate_product_values(
        name=str(payload.name or ""),
        description=str(payload.description or ""),
        base_price=base_price_value,
        sku=(payload.sku or "").strip() or None,
        color=payload.color,
        size=payload.size,
        stock=payload.stock,
    )

    p = Product(
        store_id=store_id,
        category_id=int(payload.category_id),
        name=str(payload.name).strip(),
        description=str(payload.description or ""),
        image_url=str(payload.image_url or ""),
        base_price=base_price_value,
        is_active=bool(payload.is_active),
    )
    db.add(p)
    db.flush()
    _sync_product_variants(
        db,
        store_id=store_id,
        product=p,
        colors_raw=payload.color,
        sizes_raw=payload.size,
        stock=payload.stock,
        sku=payload.sku,
    )
    db.commit()
    db.refresh(p)
    write_audit_log(
        db,
        request=request,
        user=user,
        store_id=store_id,
        action="product.create",
        entity_type="product",
        entity_id=p.id,
        before_data=None,
        after_data={
            "product": _snapshot_product(p),
            "variant_primary": _snapshot_primary_variant(db, store_id=store_id, product_id=p.id),
        },
    )
    db.commit()
    return {"id": p.id}


@router.patch("/products/{product_id}", response_model=ProductAdminOut)
def update_product(
    store_id: int,
    product_id: int,
    payload: ProductUpdateIn,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
    user: UserOut = Depends(get_current_user),
):
    product = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    before_product = _snapshot_product(product)
    before_variant = _snapshot_primary_variant(db, store_id=store_id, product_id=product_id)
    fields_set = payload.model_fields_set

    if "category_id" in fields_set:
        if payload.category_id is None:
            raise HTTPException(status_code=400, detail="category_id cannot be null")
        _ensure_store_category(db, store_id=store_id, category_id=payload.category_id)
        product.category_id = payload.category_id

    if "name" in fields_set and payload.name is not None:
        product.name = payload.name
    elif "title" in fields_set and payload.title is not None:
        product.name = payload.title

    if "description" in fields_set:
        product.description = payload.description or ""

    if "price" in fields_set:
        if payload.price is None:
            raise HTTPException(status_code=400, detail="price cannot be null")
        price_value = Decimal(str(payload.price))
        if price_value <= 0:
            raise HTTPException(status_code=400, detail="price must be greater than 0")
        product.base_price = price_value

    if "is_active" in fields_set:
        if payload.is_active is None:
            raise HTTPException(status_code=400, detail="is_active cannot be null")
        product.is_active = payload.is_active

    if "image_url" in fields_set:
        product.image_url = payload.image_url or ""

    candidate_price = Decimal(str(product.base_price))
    candidate_stock = payload.stock if "stock" in fields_set else None
    _validate_product_values(
        name=product.name,
        description=product.description or "",
        base_price=candidate_price,
        sku=(payload.sku or "").strip() if "sku" in fields_set else None,
        color=payload.color if "color" in fields_set else None,
        size=payload.size if "size" in fields_set else None,
        stock=candidate_stock,
    )

    should_sync_variant = any(k in fields_set for k in ("price", "color", "size", "stock", "sku"))
    if should_sync_variant:
        _sync_product_variants(
            db,
            store_id=store_id,
            product=product,
            colors_raw=payload.color,
            sizes_raw=payload.size,
            stock=payload.stock,
            sku=payload.sku,
            sync_all_prices="price" in fields_set,
        )

    db.add(product)
    db.commit()
    db.refresh(product)
    write_audit_log(
        db,
        request=request,
        user=user,
        store_id=store_id,
        action="product.update",
        entity_type="product",
        entity_id=product.id,
        before_data={"product": before_product, "variant_primary": before_variant},
        after_data={
            "product": _snapshot_product(product),
            "variant_primary": _snapshot_primary_variant(db, store_id=store_id, product_id=product.id),
        },
    )
    db.commit()
    primary_variant = (
        db.query(ProductVariant)
        .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product.id)
        .order_by(ProductVariant.id.asc())
        .first()
    )
    return ProductAdminOut(
        id=product.id,
        category_id=product.category_id,
        name=product.name,
        description=product.description or "",
        image_url=product.image_url,
        images=[ProductImageAdminOut.model_validate(image) for image in product.images],
        base_price=product.base_price,
        is_active=product.is_active,
        sku=primary_variant.sku if primary_variant else "",
        color=primary_variant.color if primary_variant else "",
        size=primary_variant.size if primary_variant else "",
        stock=int(primary_variant.stock) if primary_variant else 0,
    )


@router.delete("/products/{product_id}", response_model=ProductDeleteOut)
def delete_product(
    store_id: int,
    product_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
    user: UserOut = Depends(get_current_user),
):
    """Delete product with safety rule.

    - Soft-delete (is_active=false) when order history exists.
    - Hard-delete otherwise.
    """
    product = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    before_product = _snapshot_product(product)
    before_variant = _snapshot_primary_variant(db, store_id=store_id, product_id=product_id)
    has_order_history = (
        db.query(OrderItem.id)
        .filter(OrderItem.store_id == store_id, OrderItem.product_id == product_id)
        .first()
        is not None
    )

    if has_order_history:
        product.is_active = False
        db.add(product)
        db.commit()
        write_audit_log(
            db,
            request=request,
            user=user,
            store_id=store_id,
            action="product.soft_delete",
            entity_type="product",
            entity_id=product.id,
            before_data={"product": before_product, "variant_primary": before_variant},
            after_data={
                "product": _snapshot_product(product),
                "variant_primary": _snapshot_primary_variant(db, store_id=store_id, product_id=product.id),
            },
        )
        db.commit()
        return {"ok": True, "mode": "soft_delete", "message": "Product has order history and was deactivated"}

    db.delete(product)
    db.commit()
    write_audit_log(
        db,
        request=request,
        user=user,
        store_id=store_id,
        action="product.hard_delete",
        entity_type="product",
        entity_id=product_id,
        before_data={"product": before_product, "variant_primary": before_variant},
        after_data=None,
    )
    db.commit()
    return {"ok": True, "mode": "hard_delete", "message": "Product deleted permanently"}


@router.post("/products/bulk-update", response_model=ProductBulkUpdateOut)
def bulk_update_products(
    store_id: int,
    payload: ProductBulkUpdateIn,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
    user: UserOut = Depends(get_current_user),
):
    product_ids = [int(pid) for pid in payload.product_ids if int(pid) > 0]
    if not product_ids:
        raise HTTPException(status_code=400, detail="product_ids required")

    if payload.category_id is not None:
        _ensure_store_category(db, store_id=store_id, category_id=int(payload.category_id))
    if payload.price is not None and Decimal(str(payload.price)) <= 0:
        raise HTTPException(status_code=400, detail="price must be greater than 0")
    if payload.stock is not None and int(payload.stock) < 0:
        raise HTTPException(status_code=400, detail="stock must be greater than or equal to 0")

    rows = db.query(Product).filter(Product.store_id == store_id, Product.id.in_(product_ids)).all()
    updated_count = 0
    for product in rows:
        if payload.category_id is not None:
            product.category_id = int(payload.category_id)
        if payload.is_active is not None:
            product.is_active = bool(payload.is_active)
        if payload.price is not None:
            product.base_price = Decimal(str(payload.price))

        if payload.price is not None or payload.stock is not None:
            primary = (
                db.query(ProductVariant)
                .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == product.id)
                .order_by(ProductVariant.id.asc())
                .first()
            )
            if primary:
                if payload.price is not None:
                    primary.price = Decimal(str(payload.price))
                if payload.stock is not None:
                    primary.stock = int(payload.stock)
                db.add(primary)
        db.add(product)
        updated_count += 1

    db.commit()
    write_audit_log(
        db,
        request=request,
        user=user,
        store_id=store_id,
        action="catalog.bulk_update",
        entity_type="product",
        entity_id=None,
        before_data={"product_ids": product_ids},
        after_data={
            "updated_count": updated_count,
            "category_id": payload.category_id,
            "is_active": payload.is_active,
            "price": float(payload.price) if payload.price is not None else None,
            "stock": payload.stock,
        },
    )
    db.commit()
    return ProductBulkUpdateOut(ok=True, updated_count=updated_count)


@router.post("/products/{product_id}/duplicate", response_model=ProductAdminOut)
def duplicate_product(
    store_id: int,
    product_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
    user: UserOut = Depends(get_current_user),
):
    source = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    source_variants = (
        db.query(ProductVariant)
        .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == source.id)
        .order_by(ProductVariant.id.asc())
        .all()
    )
    source_images = (
        db.query(ProductImage)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == source.id)
        .order_by(ProductImage.sort_order.asc(), ProductImage.id.asc())
        .all()
    )

    clone = Product(
        store_id=store_id,
        category_id=source.category_id,
        name=f"{source.name} (Copia)",
        description=source.description or "",
        image_url=source.image_url,
        base_price=source.base_price,
        is_active=False,
    )
    db.add(clone)
    db.flush()

    for variant in source_variants:
        new_variant = ProductVariant(
            store_id=store_id,
            product_id=clone.id,
            sku=_build_unique_sku(db, store_id=store_id, candidate=f"{variant.sku}-COPY"),
            color=variant.color,
            size=variant.size,
            price=variant.price,
            stock=variant.stock,
            active=False,
        )
        db.add(new_variant)

    for image in source_images:
        db.add(
            ProductImage(
                store_id=store_id,
                product_id=clone.id,
                image_url=image.image_url,
                sort_order=image.sort_order,
                is_cover=image.is_cover,
            )
        )

    db.commit()
    db.refresh(clone)
    primary_variant = (
        db.query(ProductVariant)
        .filter(ProductVariant.store_id == store_id, ProductVariant.product_id == clone.id)
        .order_by(ProductVariant.id.asc())
        .first()
    )
    write_audit_log(
        db,
        request=request,
        user=user,
        store_id=store_id,
        action="product.duplicate",
        entity_type="product",
        entity_id=clone.id,
        before_data={"source_product_id": source.id},
        after_data={"new_product_id": clone.id},
    )
    db.commit()
    return ProductAdminOut(
        id=clone.id,
        category_id=clone.category_id,
        name=clone.name,
        description=clone.description or "",
        image_url=clone.image_url,
        images=[ProductImageAdminOut.model_validate(image) for image in clone.images],
        base_price=clone.base_price,
        is_active=clone.is_active,
        sku=primary_variant.sku if primary_variant else "",
        color=primary_variant.color if primary_variant else "",
        size=primary_variant.size if primary_variant else "",
        stock=int(primary_variant.stock) if primary_variant else 0,
    )


@router.get("/catalog-jobs", response_model=list[CatalogJobOut])
def list_catalog_jobs(
    store_id: int,
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    rows = (
        db.query(CatalogJob)
        .filter(CatalogJob.store_id == store_id)
        .order_by(CatalogJob.created_at.desc(), CatalogJob.id.desc())
        .limit(limit)
        .all()
    )
    return [CatalogJobOut.model_validate(row) for row in rows]


@router.get("/catalog-jobs/{job_id}", response_model=CatalogJobOut)
def get_catalog_job(
    store_id: int,
    job_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    row = db.query(CatalogJob).filter(CatalogJob.store_id == store_id, CatalogJob.id == job_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="catalog job not found")
    return CatalogJobOut.model_validate(row)


@router.post("/catalog-jobs/reprocess-images", response_model=CatalogJobOut)
def enqueue_reprocess_images(
    store_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
    user: UserOut = Depends(get_current_user),
):
    job = CatalogJob(
        store_id=store_id,
        user_id=user.id,
        job_type="image_reprocess",
        status="completed",
        payload={"scope": "all_products"},
        result={"message": "placeholder queue created; processor can be attached later"},
        started_at=datetime.utcnow(),
        finished_at=datetime.utcnow(),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    write_audit_log(
        db,
        request=request,
        user=user,
        store_id=store_id,
        action="catalog.image_reprocess.enqueue",
        entity_type="catalog_job",
        entity_id=job.id,
        before_data=None,
        after_data=job.result or {},
    )
    db.commit()
    return CatalogJobOut.model_validate(job)


@router.post("/products/{product_id}/variants")
def add_variant(
    store_id: int,
    product_id: int,
    payload: ProductVariantCreateIn,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    _ensure_store_product(db, store_id=store_id, product_id=product_id)

    required = ["sku", "price", "stock"]
    for k in required:
        if getattr(payload, k) in (None, ""):
            raise HTTPException(status_code=400, detail=f"{k} required")

    price_value = Decimal(str(payload.price))

    v = ProductVariant(
        store_id=store_id,
        product_id=product_id,
        sku=str(payload.sku),
        color=str(payload.color or ""),
        size=str(payload.size or ""),
        price=price_value,
        stock=int(payload.stock),
        active=bool(payload.active),
    )
    db.add(v)
    db.commit()
    db.refresh(v)
    return {"id": v.id}


@router.get("/products/{product_id}/images", response_model=list[ProductImageAdminOut])
def list_product_images(
    store_id: int,
    product_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    _ensure_store_product(db, store_id=store_id, product_id=product_id)
    rows = (
        db.query(ProductImage)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id)
        .order_by(ProductImage.sort_order.asc(), ProductImage.id.asc())
        .all()
    )
    return [ProductImageAdminOut.model_validate(row) for row in rows]


@router.post("/products/{product_id}/images", response_model=ProductImageAdminOut)
async def upload_product_gallery_image(
    store_id: int,
    product_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    product = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    try:
        saved = await save_upload_file(file, target_dir=get_product_upload_dir(product_id))
    except UploadValidationError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    existing_count = (
        db.query(ProductImage.id)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id)
        .count()
    )
    image = ProductImage(
        store_id=store_id,
        product_id=product_id,
        image_url=saved.public_url,
        sort_order=existing_count,
        is_cover=existing_count == 0,
    )
    db.add(image)

    if image.is_cover:
        product.image_url = saved.public_url
        db.add(product)

    db.commit()
    db.refresh(image)
    return ProductImageAdminOut.model_validate(image)


@router.patch("/products/{product_id}/images/{image_id}", response_model=ProductImageAdminOut)
def update_product_gallery_image(
    store_id: int,
    product_id: int,
    image_id: int,
    payload: ProductImageUpdateIn,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    product = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    image = _ensure_store_product_image(db, store_id=store_id, product_id=product_id, image_id=image_id)
    fields_set = payload.model_fields_set

    if "sort_order" in fields_set and payload.sort_order is not None:
        image.sort_order = payload.sort_order

    if "is_cover" in fields_set and payload.is_cover is not None:
        if payload.is_cover:
            _set_cover_image(db, store_id=store_id, product_id=product_id, image_id=image_id)
        else:
            image.is_cover = False
            db.add(image)

    db.add(image)
    _sync_product_cover_from_gallery(db, store_id=store_id, product=product)
    db.commit()
    db.refresh(image)
    return ProductImageAdminOut.model_validate(image)


@router.delete("/products/{product_id}/images/{image_id}", response_model=list[ProductImageAdminOut])
def delete_product_gallery_image(
    store_id: int,
    product_id: int,
    image_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    product = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    image = _ensure_store_product_image(db, store_id=store_id, product_id=product_id, image_id=image_id)

    delete_upload_by_public_url(image.image_url)
    db.delete(image)
    db.flush()

    rows = (
        db.query(ProductImage)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id)
        .order_by(ProductImage.sort_order.asc(), ProductImage.id.asc())
        .all()
    )
    for idx, row in enumerate(rows):
        row.sort_order = idx
        db.add(row)

    _sync_product_cover_from_gallery(db, store_id=store_id, product=product)
    db.commit()
    return [ProductImageAdminOut.model_validate(row) for row in rows]


@router.post("/products/{product_id}/image", response_model=ProductImageOut)
async def upload_product_image(
    store_id: int,
    product_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    """Upload product cover image.

    multipart/form-data:
    - file: <binary image>
    """
    product = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    try:
        saved = await save_upload_file(file, target_dir=get_product_upload_dir(product_id))
    except UploadValidationError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    previous_cover = (
        db.query(ProductImage)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id, ProductImage.is_cover == True)  # noqa: E712
        .first()
    )
    if previous_cover:
        delete_upload_by_public_url(previous_cover.image_url)
        previous_cover.image_url = saved.public_url
        db.add(previous_cover)
    else:
        next_order = (
            db.query(ProductImage.id)
            .filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id)
            .count()
        )
        cover = ProductImage(
            store_id=store_id,
            product_id=product_id,
            image_url=saved.public_url,
            sort_order=next_order,
            is_cover=True,
        )
        db.add(cover)

    product.image_url = saved.public_url
    db.add(product)
    db.commit()
    return ProductImageOut(product_id=product.id, image_url=product.image_url)


@router.delete("/products/{product_id}/image", response_model=ProductImageOut)
def delete_product_image(
    store_id: int,
    product_id: int,
    delete_file: bool = Query(default=True),
    db: Session = Depends(get_db),
    _=Depends(require_catalog_write),
):
    product = _ensure_store_product(db, store_id=store_id, product_id=product_id)
    cover = (
        db.query(ProductImage)
        .filter(ProductImage.store_id == store_id, ProductImage.product_id == product_id, ProductImage.is_cover == True)  # noqa: E712
        .first()
    )

    if cover:
        cover.is_cover = False
        db.add(cover)
        if delete_file:
            delete_upload_by_public_url(cover.image_url)
            db.delete(cover)

    _sync_product_cover_from_gallery(db, store_id=store_id, product=product)
    db.commit()
    db.refresh(product)
    return ProductImageOut(product_id=product.id, image_url=product.image_url)
